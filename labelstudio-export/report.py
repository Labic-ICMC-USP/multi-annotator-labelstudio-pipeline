
from __future__ import annotations

import argparse
import json
import math
from collections import Counter
from datetime import datetime, timezone
from itertools import combinations
from pathlib import Path
from typing import Any, Optional, Sequence, Tuple

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    from artifact_tool import SpreadsheetArtifact
except Exception:
    SpreadsheetArtifact = None  # type: ignore

SPAN = Tuple[str, int, int]


def infer_annotator_email(project_title: str, task: dict[str, Any]) -> str:
    data = task.get("data", {}) or {}
    if data.get("assigned_annotator_email"):
        return str(data["assigned_annotator_email"]).strip()
    title = project_title or ""
    for sep in ("__", " :: "):
        if sep in title:
            return title.split(sep, 1)[1].strip()
    return "unknown"


def parse_dt(value: Any) -> str:
    return "" if not value else str(value)


def pick_annotation(task: dict[str, Any]) -> Optional[dict[str, Any]]:
    anns = [a for a in (task.get("annotations") or []) if not a.get("was_cancelled", False)]
    if not anns:
        return None
    anns.sort(key=lambda a: (a.get("updated_at") or "", a.get("created_at") or "", a.get("id") or 0))
    return anns[-1]


def extract_spans(annotation: Optional[dict[str, Any]]) -> list[dict[str, Any]]:
    if not annotation:
        return []
    spans: list[dict[str, Any]] = []
    for idx, result in enumerate(annotation.get("result") or [], start=1):
        value = result.get("value") or {}
        labels = value.get("labels") or []
        if not labels:
            continue
        start = value.get("start")
        end = value.get("end")
        text = value.get("text", "")
        if start is None or end is None:
            continue
        for label in labels:
            spans.append(
                {
                    "span_idx": idx,
                    "label": str(label),
                    "start": int(start),
                    "end": int(end),
                    "text": str(text),
                    "length": int(end) - int(start),
                    "origin": result.get("origin", ""),
                    "from_name": result.get("from_name", ""),
                    "to_name": result.get("to_name", ""),
                }
            )
    spans.sort(key=lambda s: (s["start"], s["end"], s["label"], s["text"]))
    return spans


def span_key(span: dict[str, Any]) -> SPAN:
    return (str(span["label"]), int(span["start"]), int(span["end"]))


def spans_to_string(spans: Sequence[dict[str, Any]]) -> str:
    if not spans:
        return ""
    parts: list[str] = []
    for sp in spans:
        txt = str(sp["text"]).replace("\n", " ").strip()
        if len(txt) > 50:
            txt = txt[:47] + "..."
        parts.append(f'{sp["label"]}[{sp["start"]}:{sp["end"]}]="{txt}"')
    return " ; ".join(parts)


def overlap(a: SPAN, b: SPAN) -> bool:
    return max(a[1], b[1]) < min(a[2], b[2])


def categorize_span_differences(a_only: set[SPAN], b_only: set[SPAN]) -> tuple[int, int, int, list[str]]:
    remaining_b = set(b_only)
    boundary = 0
    label = 0
    examples: list[str] = []

    for a in sorted(a_only):
        candidate_same = None
        candidate_diff = None
        for b in sorted(remaining_b):
            if overlap(a, b):
                if a[0] == b[0]:
                    candidate_same = b
                    break
                if candidate_diff is None:
                    candidate_diff = b
        if candidate_same is not None:
            boundary += 1
            remaining_b.remove(candidate_same)
            if len(examples) < 4:
                examples.append(f"boundary: {a} vs {candidate_same}")
        elif candidate_diff is not None:
            label += 1
            remaining_b.remove(candidate_diff)
            if len(examples) < 4:
                examples.append(f"label: {a} vs {candidate_diff}")

    missing_extra = max(0, len(a_only) + len(b_only) - 2 * (boundary + label))
    return boundary, label, missing_extra, examples


def build_char_labels(text: str, spans: Sequence[dict[str, Any]]) -> list[str]:
    seq = ["O"] * len(text)
    for sp in spans:
        label = str(sp["label"])
        start = max(0, min(len(text), int(sp["start"])))
        end = max(0, min(len(text), int(sp["end"])))
        for i in range(start, end):
            if seq[i] == "O":
                seq[i] = label
            elif seq[i] != label:
                seq[i] = "OVERLAP"
    return seq


def cohen_kappa_from_sequences(seq_a: Sequence[str], seq_b: Sequence[str]) -> tuple[Optional[float], Optional[float]]:
    n = min(len(seq_a), len(seq_b))
    if n == 0:
        return None, None
    seq_a = seq_a[:n]
    seq_b = seq_b[:n]
    agree = sum(1 for x, y in zip(seq_a, seq_b) if x == y)
    po = agree / n
    counts_a = Counter(seq_a)
    counts_b = Counter(seq_b)
    labels = set(counts_a) | set(counts_b)
    pe = sum((counts_a[l] / n) * (counts_b[l] / n) for l in labels)
    if abs(1 - pe) < 1e-12:
        kappa = None
    else:
        kappa = (po - pe) / (1 - pe)
    return po, kappa


def binary_kappa(seq_a: Sequence[str], seq_b: Sequence[str], positive_label: str) -> tuple[Optional[float], Optional[float], float, float]:
    if not seq_a or not seq_b:
        return None, None, 0.0, 0.0
    a = ["POS" if x == positive_label else "NEG" for x in seq_a]
    b = ["POS" if x == positive_label else "NEG" for x in seq_b]
    po, kappa = cohen_kappa_from_sequences(a, b)
    pos_a = a.count("POS") / len(a)
    pos_b = b.count("POS") / len(b)
    return po, kappa, pos_a, pos_b


def avg(values: Sequence[Optional[float]]) -> Optional[float]:
    vals = [float(v) for v in values if v is not None]
    return None if not vals else sum(vals) / len(vals)


def analyze_export(payload: dict[str, Any]) -> dict[str, Any]:
    grouped: dict[str, dict[str, Any]] = {}
    spans_rows: list[dict[str, Any]] = []
    project_meta_rows: list[dict[str, Any]] = []
    global_annotators: set[str] = set()
    global_labels: set[str] = set()

    for project_entry in payload.get("projects", []):
        project = project_entry.get("project", {}) or {}
        project_id = project.get("id")
        project_title = str(project.get("title", ""))
        tasks = project_entry.get("tasks") or []
        project_meta_rows.append(
            {
                "project_id": project_id,
                "project_title": project_title,
                "task_number": project.get("task_number", 0),
                "num_tasks_with_annotations": project.get("num_tasks_with_annotations", 0),
                "total_annotations_number": project.get("total_annotations_number", 0),
                "exported_tasks": len(tasks),
            }
        )

        for task in tasks:
            data = task.get("data", {}) or {}
            external_id = str(data.get("external_id") or f'task_{task.get("id")}')
            project_code = str(data.get("project_code") or "")
            group_key = f"{project_code}::{external_id}" if project_code else external_id
            annotator_email = infer_annotator_email(project_title, task)
            global_annotators.add(annotator_email)

            selected_annotation = pick_annotation(task)
            spans = extract_spans(selected_annotation)
            for sp in spans:
                global_labels.add(str(sp["label"]))

            group = grouped.setdefault(
                group_key,
                {
                    "group_key": group_key,
                    "external_id": external_id,
                    "project_code": project_code,
                    "project_title": str(data.get("project_title") or ""),
                    "batch": str(data.get("batch") or ""),
                    "source": str(data.get("source") or ""),
                    "text": str(data.get("text") or ""),
                    "text_length": len(str(data.get("text") or "")),
                    "annotators": {},
                },
            )

            group["annotators"][annotator_email] = {
                "annotator_email": annotator_email,
                "project_id": project_id,
                "project_title": project_title,
                "task_id": task.get("id"),
                "annotation_id": selected_annotation.get("id") if selected_annotation else None,
                "completed_by": selected_annotation.get("completed_by") if selected_annotation else None,
                "annotation_created_at": parse_dt(selected_annotation.get("created_at") if selected_annotation else ""),
                "annotation_updated_at": parse_dt(selected_annotation.get("updated_at") if selected_annotation else ""),
                "lead_time": selected_annotation.get("lead_time") if selected_annotation else None,
                "spans": spans,
                "task_data": data,
            }

            for sp in spans:
                spans_rows.append(
                    {
                        "group_key": group_key,
                        "external_id": external_id,
                        "project_code": project_code,
                        "project_title": str(data.get("project_title") or ""),
                        "batch": str(data.get("batch") or ""),
                        "source": str(data.get("source") or ""),
                        "annotator_email": annotator_email,
                        "project_id": project_id,
                        "project_name": project_title,
                        "task_id": task.get("id"),
                        "annotation_id": selected_annotation.get("id") if selected_annotation else None,
                        "completed_by": selected_annotation.get("completed_by") if selected_annotation else None,
                        "annotation_created_at": parse_dt(selected_annotation.get("created_at") if selected_annotation else ""),
                        "annotation_updated_at": parse_dt(selected_annotation.get("updated_at") if selected_annotation else ""),
                        "lead_time": selected_annotation.get("lead_time") if selected_annotation else None,
                        "span_idx": sp["span_idx"],
                        "label": sp["label"],
                        "start": sp["start"],
                        "end": sp["end"],
                        "length": sp["length"],
                        "text": sp["text"],
                        "origin": sp["origin"],
                        "from_name": sp["from_name"],
                        "to_name": sp["to_name"],
                    }
                )

    annotators_sorted = sorted(global_annotators)
    labels_sorted = sorted(global_labels)

    grouped_rows: list[dict[str, Any]] = []
    task_agreement_rows: list[dict[str, Any]] = []
    pairwise_rows: list[dict[str, Any]] = []
    label_count_rows: list[dict[str, Any]] = []
    label_kappa_rows: list[dict[str, Any]] = []

    for group_key, group in sorted(grouped.items()):
        annot_map = group["annotators"]
        annotator_names = sorted(annot_map)
        span_strings = {a: spans_to_string(annot_map[a]["spans"]) for a in annotator_names}
        label_union = sorted({sp["label"] for a in annotator_names for sp in annot_map[a]["spans"]})

        grouped_row = {
            "group_key": group_key,
            "external_id": group["external_id"],
            "project_code": group["project_code"],
            "project_title": group["project_title"],
            "batch": group["batch"],
            "source": group["source"],
            "text": group["text"],
            "text_length": group["text_length"],
            "annotators": ", ".join(annotator_names),
            "n_annotators": len(annotator_names),
            "task_ids": ", ".join(str(annot_map[a]["task_id"]) for a in annotator_names),
            "project_ids": ", ".join(str(annot_map[a]["project_id"]) for a in annotator_names),
            "annotation_ids": ", ".join(str(annot_map[a]["annotation_id"]) for a in annotator_names if annot_map[a]["annotation_id"] is not None),
            "total_spans": sum(len(annot_map[a]["spans"]) for a in annotator_names),
            "labels_union": ", ".join(label_union),
        }
        for annotator in annotators_sorted:
            grouped_row[f"annotator::{annotator}"] = span_strings.get(annotator, "")
        grouped_rows.append(grouped_row)

        for annotator in annotator_names:
            counts = Counter(sp["label"] for sp in annot_map[annotator]["spans"])
            for label in labels_sorted:
                label_count_rows.append(
                    {
                        "annotator_email": annotator,
                        "label": label,
                        "span_count": counts.get(label, 0),
                        "group_key": group_key,
                        "external_id": group["external_id"],
                        "project_code": group["project_code"],
                    }
                )

        if len(annotator_names) < 2:
            task_agreement_rows.append(
                {
                    "group_key": group_key,
                    "external_id": group["external_id"],
                    "project_code": group["project_code"],
                    "project_title": group["project_title"],
                    "batch": group["batch"],
                    "source": group["source"],
                    "n_annotators": len(annotator_names),
                    "annotators": ", ".join(annotator_names),
                    "compared": 0,
                    "pair_count": 0,
                    "exact_agreement_all": None,
                    "avg_pair_jaccard": None,
                    "avg_pair_span_precision": None,
                    "avg_pair_span_recall": None,
                    "avg_pair_span_f1": None,
                    "avg_pair_char_agreement": None,
                    "avg_pair_char_kappa": None,
                    "shared_exact_spans_avg": None,
                    "union_exact_spans_avg": None,
                    "exclusive_exact_spans_avg": None,
                    "boundary_mismatch_avg": None,
                    "label_mismatch_avg": None,
                    "missing_extra_avg": None,
                    "disagreement_examples": "Only one annotator available for this external_id.",
                }
            )
            continue

        pair_jaccards: list[Optional[float]] = []
        pair_precisions: list[Optional[float]] = []
        pair_recalls: list[Optional[float]] = []
        pair_f1s: list[Optional[float]] = []
        pair_char_agreements: list[Optional[float]] = []
        pair_char_kappas: list[Optional[float]] = []
        shared_counts: list[int] = []
        union_counts: list[int] = []
        exclusive_counts: list[int] = []
        boundary_counts: list[int] = []
        label_counts: list[int] = []
        missing_counts: list[int] = []
        all_examples: list[str] = []
        exact_all = True
        text = group["text"]

        for annotator_a, annotator_b in combinations(annotator_names, 2):
            spans_a = annot_map[annotator_a]["spans"]
            spans_b = annot_map[annotator_b]["spans"]
            set_a = {span_key(sp) for sp in spans_a}
            set_b = {span_key(sp) for sp in spans_b}
            shared = set_a & set_b
            union = set_a | set_b
            a_only = set_a - set_b
            b_only = set_b - set_a

            precision = len(shared) / len(set_a) if set_a else (1.0 if not set_b else 0.0)
            recall = len(shared) / len(set_b) if set_b else (1.0 if not set_a else 0.0)
            f1 = 0.0 if precision + recall == 0 else 2 * precision * recall / (precision + recall)
            jaccard = len(shared) / len(union) if union else 1.0
            exact_pair = set_a == set_b
            if not exact_pair:
                exact_all = False

            boundary, label_mismatch, missing_extra, examples = categorize_span_differences(a_only, b_only)
            all_examples.extend(examples)

            seq_a = build_char_labels(text, spans_a)
            seq_b = build_char_labels(text, spans_b)
            char_agreement, char_kappa = cohen_kappa_from_sequences(seq_a, seq_b)

            pair_jaccards.append(jaccard)
            pair_precisions.append(precision)
            pair_recalls.append(recall)
            pair_f1s.append(f1)
            pair_char_agreements.append(char_agreement)
            pair_char_kappas.append(char_kappa)
            shared_counts.append(len(shared))
            union_counts.append(len(union))
            exclusive_counts.append(len(union) - len(shared))
            boundary_counts.append(boundary)
            label_counts.append(label_mismatch)
            missing_counts.append(missing_extra)

            pairwise_rows.append(
                {
                    "group_key": group_key,
                    "external_id": group["external_id"],
                    "project_code": group["project_code"],
                    "project_title": group["project_title"],
                    "annotator_a": annotator_a,
                    "annotator_b": annotator_b,
                    "task_id_a": annot_map[annotator_a]["task_id"],
                    "task_id_b": annot_map[annotator_b]["task_id"],
                    "annotation_id_a": annot_map[annotator_a]["annotation_id"],
                    "annotation_id_b": annot_map[annotator_b]["annotation_id"],
                    "spans_a": len(set_a),
                    "spans_b": len(set_b),
                    "shared_exact_spans": len(shared),
                    "union_exact_spans": len(union),
                    "jaccard_exact_spans": jaccard,
                    "precision_a_vs_b": precision,
                    "recall_a_vs_b": recall,
                    "f1_exact_spans": f1,
                    "text_length_chars": len(text),
                    "char_agreement": char_agreement,
                    "char_kappa": char_kappa,
                    "boundary_mismatch_count": boundary,
                    "label_mismatch_count": label_mismatch,
                    "missing_extra_count": missing_extra,
                    "a_only_spans": " ; ".join(map(str, sorted(a_only))),
                    "b_only_spans": " ; ".join(map(str, sorted(b_only))),
                    "examples": " | ".join(examples[:4]),
                }
            )

            for label in labels_sorted:
                po, kap, pos_a, pos_b = binary_kappa(seq_a, seq_b, label)
                label_kappa_rows.append(
                    {
                        "group_key": group_key,
                        "external_id": group["external_id"],
                        "project_code": group["project_code"],
                        "annotator_a": annotator_a,
                        "annotator_b": annotator_b,
                        "label": label,
                        "char_agreement_one_vs_rest": po,
                        "char_kappa_one_vs_rest": kap,
                        "positive_rate_a": pos_a,
                        "positive_rate_b": pos_b,
                    }
                )

        task_agreement_rows.append(
            {
                "group_key": group_key,
                "external_id": group["external_id"],
                "project_code": group["project_code"],
                "project_title": group["project_title"],
                "batch": group["batch"],
                "source": group["source"],
                "n_annotators": len(annotator_names),
                "annotators": ", ".join(annotator_names),
                "compared": 1,
                "pair_count": len(pair_jaccards),
                "exact_agreement_all": 1 if exact_all else 0,
                "avg_pair_jaccard": avg(pair_jaccards),
                "avg_pair_span_precision": avg(pair_precisions),
                "avg_pair_span_recall": avg(pair_recalls),
                "avg_pair_span_f1": avg(pair_f1s),
                "avg_pair_char_agreement": avg(pair_char_agreements),
                "avg_pair_char_kappa": avg(pair_char_kappas),
                "shared_exact_spans_avg": avg(shared_counts),
                "union_exact_spans_avg": avg(union_counts),
                "exclusive_exact_spans_avg": avg(exclusive_counts),
                "boundary_mismatch_avg": avg(boundary_counts),
                "label_mismatch_avg": avg(label_counts),
                "missing_extra_avg": avg(missing_counts),
                "disagreement_examples": " | ".join(all_examples[:4]) if all_examples else "",
            }
        )

    return {
        "grouped_rows": grouped_rows,
        "spans_rows": spans_rows,
        "task_agreement_rows": task_agreement_rows,
        "pairwise_rows": pairwise_rows,
        "label_count_rows": label_count_rows,
        "label_kappa_rows": label_kappa_rows,
        "project_meta_rows": project_meta_rows,
        "annotators_sorted": annotators_sorted,
        "labels_sorted": labels_sorted,
        "payload": payload,
    }


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(bottom=thin)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"


def autofit(ws) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells:
            value = "" if c.value is None else str(c.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)
    ws.auto_filter.ref = ws.dimensions


def append_sheet(ws, rows: list[dict[str, Any]], columns: list[str]) -> None:
    ws.append(columns)
    for row in rows:
        ws.append([row.get(col) for col in columns])
    style_header(ws)
    autofit(ws)


def create_workbook(analysis: dict[str, Any], output_path: Path, input_path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Label Studio Consensus Report"
    summary["A1"].font = Font(bold=True, size=14)
    summary["A3"] = "Input file"
    summary["B3"] = str(input_path)
    summary["A4"] = "Generated at (UTC)"
    summary["B4"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    summary["A5"] = "Base URL"
    summary["B5"] = analysis["payload"].get("base_url", "")
    summary["A6"] = "Project filter"
    summary["B6"] = analysis["payload"].get("project_filter", "")
    summary["A8"] = "Metric"
    summary["B8"] = "Value"
    style_header(summary, 8)

    summary_rows = [
        ("Total grouped tasks", "=COUNTA(GroupedTasks!A:A)-1"),
        ("Tasks with >=2 annotators", '=COUNTIF(GroupedTasks!J:J,">=2")'),
        ("Average annotators per task", "=AVERAGE(GroupedTasks!J:J)"),
        ("Total span rows", "=COUNTA(AnnotationSpans!A:A)-1"),
        ("Pairwise comparisons", "=COUNTA(PairwiseMetrics!A:A)-1"),
        ("Comparable tasks in TaskAgreement", "=SUM(TaskAgreement!I:I)"),
        ("Exact agreement rate (comparable tasks)", '=IFERROR(AVERAGEIF(TaskAgreement!K:K,"<>",TaskAgreement!K:K),"")'),
        ("Average pairwise Jaccard", '=IFERROR(AVERAGEIF(PairwiseMetrics!O:O,"<>",PairwiseMetrics!O:O),"")'),
        ("Average pairwise span F1", '=IFERROR(AVERAGEIF(PairwiseMetrics!R:R,"<>",PairwiseMetrics!R:R),"")'),
        ("Average pairwise char agreement", '=IFERROR(AVERAGEIF(PairwiseMetrics!T:T,"<>",PairwiseMetrics!T:T),"")'),
        ("Average pairwise char kappa", '=IFERROR(AVERAGEIF(PairwiseMetrics!U:U,"<>",PairwiseMetrics!U:U),"")'),
    ]
    start = 9
    for idx, (label, formula) in enumerate(summary_rows, start=start):
        summary[f"A{idx}"] = label
        summary[f"B{idx}"] = formula

    notes_row = start + len(summary_rows) + 2
    summary[f"A{notes_row}"] = "Notes"
    summary[f"A{notes_row}"].font = Font(bold=True)
    notes = [
        "GroupedTasks: one row per external_id (grouped with project_code when available).",
        "AnnotationSpans: one row per labeled span.",
        "TaskAgreement: average agreement metrics across annotator pairs for each grouped task.",
        "PairwiseMetrics: exact-span agreement and char-level agreement for each annotator pair.",
        "PairwiseLabelKappa: one-vs-rest kappa per label using character positions.",
        "Character-level kappa uses raw character positions with labels plus O (outside).",
        "Exact span agreement requires exact match of label, start, and end offsets.",
    ]
    for i, note in enumerate(notes, start=notes_row + 1):
        summary[f"A{i}"] = f"- {note}"
        summary.merge_cells(start_row=i, start_column=1, end_row=i, end_column=5)
        summary[f"A{i}"].alignment = Alignment(wrap_text=True)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 24

    grouped_cols = [
        "group_key", "external_id", "project_code", "project_title", "batch", "source",
        "text", "text_length", "annotators", "n_annotators", "task_ids", "project_ids",
        "annotation_ids", "total_spans", "labels_union"
    ] + [f"annotator::{a}" for a in analysis["annotators_sorted"]]
    grouped_ws = wb.create_sheet("GroupedTasks")
    append_sheet(grouped_ws, analysis["grouped_rows"], grouped_cols)
    grouped_ws.column_dimensions["G"].width = 80

    spans_cols = [
        "group_key", "external_id", "project_code", "project_title", "batch", "source",
        "annotator_email", "project_id", "project_name", "task_id", "annotation_id", "completed_by",
        "annotation_created_at", "annotation_updated_at", "lead_time", "span_idx", "label",
        "start", "end", "length", "text", "origin", "from_name", "to_name"
    ]
    spans_ws = wb.create_sheet("AnnotationSpans")
    append_sheet(spans_ws, analysis["spans_rows"], spans_cols)
    spans_ws.column_dimensions["U"].width = 60

    agreement_cols = [
        "group_key", "external_id", "project_code", "project_title", "batch", "source",
        "n_annotators", "annotators", "compared", "pair_count", "exact_agreement_all",
        "avg_pair_jaccard", "avg_pair_span_precision", "avg_pair_span_recall", "avg_pair_span_f1",
        "avg_pair_char_agreement", "avg_pair_char_kappa", "shared_exact_spans_avg",
        "union_exact_spans_avg", "exclusive_exact_spans_avg", "boundary_mismatch_avg",
        "label_mismatch_avg", "missing_extra_avg", "disagreement_examples"
    ]
    agreement_ws = wb.create_sheet("TaskAgreement")
    append_sheet(agreement_ws, analysis["task_agreement_rows"], agreement_cols)
    agreement_ws.column_dimensions["X"].width = 90

    pairwise_cols = [
        "group_key", "external_id", "project_code", "project_title", "annotator_a", "annotator_b",
        "task_id_a", "task_id_b", "annotation_id_a", "annotation_id_b", "spans_a", "spans_b",
        "shared_exact_spans", "union_exact_spans", "jaccard_exact_spans", "precision_a_vs_b",
        "recall_a_vs_b", "f1_exact_spans", "text_length_chars", "char_agreement", "char_kappa",
        "boundary_mismatch_count", "label_mismatch_count", "missing_extra_count", "a_only_spans",
        "b_only_spans", "examples"
    ]
    pairwise_ws = wb.create_sheet("PairwiseMetrics")
    append_sheet(pairwise_ws, analysis["pairwise_rows"], pairwise_cols)
    pairwise_ws.column_dimensions["Y"].width = 60
    pairwise_ws.column_dimensions["Z"].width = 60
    pairwise_ws.column_dimensions["AA"].width = 90

    label_kappa_cols = [
        "group_key", "external_id", "project_code", "annotator_a", "annotator_b", "label",
        "char_agreement_one_vs_rest", "char_kappa_one_vs_rest", "positive_rate_a", "positive_rate_b"
    ]
    label_kappa_ws = wb.create_sheet("PairwiseLabelKappa")
    append_sheet(label_kappa_ws, analysis["label_kappa_rows"], label_kappa_cols)

    label_count_cols = ["annotator_email", "label", "span_count", "group_key", "external_id", "project_code"]
    label_count_ws = wb.create_sheet("LabelCounts")
    append_sheet(label_count_ws, analysis["label_count_rows"], label_count_cols)

    project_meta_cols = ["project_id", "project_title", "task_number", "num_tasks_with_annotations", "total_annotations_number", "exported_tasks"]
    project_meta_ws = wb.create_sheet("ProjectMeta")
    append_sheet(project_meta_ws, analysis["project_meta_rows"], project_meta_cols)

    percent_formats = {
        "Summary": ["B15", "B16", "B17", "B18", "B19"],
        "TaskAgreement": ["L", "M", "N", "O", "P", "Q"],
        "PairwiseMetrics": ["O", "P", "Q", "R", "T", "U"],
        "PairwiseLabelKappa": ["G", "H", "I", "J"],
    }
    for sheet_name, refs in percent_formats.items():
        ws = wb[sheet_name]
        if sheet_name == "Summary":
            for ref in refs:
                ws[ref].number_format = "0.0%"
        else:
            for col in refs:
                for row in range(2, ws.max_row + 1):
                    ws[f"{col}{row}"].number_format = "0.0%"

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    wb.save(output_path)

    if SpreadsheetArtifact is not None:
        try:
            artifact = SpreadsheetArtifact.load(str(output_path))
            artifact.calculate()
            artifact.export(str(output_path), overwrite=True)
            artifact.render()
        except Exception:
            pass


def main() -> None:
    parser = argparse.ArgumentParser(description="Analyze Label Studio JSON export and generate a consensus workbook.")
    parser.add_argument("--input-json", required=True)
    parser.add_argument("--output-xlsx", required=True)
    args = parser.parse_args()

    input_path = Path(args.input_json)
    output_path = Path(args.output_xlsx)

    payload = json.loads(input_path.read_text(encoding="utf-8"))
    analysis = analyze_export(payload)
    create_workbook(analysis, output_path, input_path)


if __name__ == "__main__":
    main()
